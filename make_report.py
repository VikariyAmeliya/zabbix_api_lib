from pathlib import Path
import json
import time
from datetime import datetime, timedelta
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from configparser import ConfigParser
from zabbix_api import API

class ZabbixEventReport:
    def __init__(self, api, config_path):
        self.api = api
        self.config = self._load_config(config_path)
        
        
        self.cache_dir = Path(self.config.get('paths', 'folder_cache', fallback='/tmp/zabbix/cache_structure'))
        self.event_cache_dir = Path(self.config.get('paths', 'folder_cache_event', fallback='/tmp/zabbix/cache_event'))
        self.pid_dir = Path(self.config.get('paths', 'folder_pid', fallback='/tmp/zabbix/pid'))
        self.report_dir = Path(self.config.get('paths', 'folder_report', fallback='/tmp/zabbix/report'))
        
        self.report_event_day = self.config.getint('settings', 'report_event_day', fallback=14)
        self.mail_server = self.config.get('mail', 'server', fallback='localhost')
        self.mail_from = self.config.get('mail', 'from', fallback='zabbix@example.com')
        
        
        self.pid_dir.mkdir(parents=True, exist_ok=True)
        self.report_dir.mkdir(parents=True, exist_ok=True)
        
        self.pid_file = self.pid_dir / "event_report.pid"
        
        # Кэш данных
        self.cache = {}
        self.cache_event = {'day': {}}
        
    def _load_config(self, config_path):
        config = ConfigParser()
        config.read(config_path)
        return config
    
    def _check_proc(self):
        if self.pid_file.exists():
            print(f"Process already running, PID file exists: {self.pid_file}")
            return True
        return False
    
    def _create_pid(self):
        with open(self.pid_file, 'w') as f:
            f.write(str(os.getpid()))
            
    def _remove_pid(self):
        if self.pid_file.exists():
            self.pid_file.unlink()
    
    def _load_cache(self):
        cache_file = self.cache_dir / "zabbix_svp.cache"
        if not cache_file.exists():
            print(f"Cache file not found: {cache_file}")
            return False
        
        try:
            with open(cache_file, 'r') as f:
                self.cache = json.load(f)
            print(f"Cache loaded successfully from {cache_file}")
            return True
        except Exception as e:
            print(f"Error loading cache: {e}")
            return False



     def _generate_report(self):
        #Содаем имя файла 
        now = datetime.now()
        report_filename = f"{now.strftime('%Y-%m-%d')}-report-event-{now.strftime('%H-%M-%S')}.xlsx"
        report_path = self.report_dir / report_filename
        
        #Новая книга Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Events Report"
        
        #Стили для заголовков
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        header_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        #Стили для ячеек
        cell_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        #Цвета для severity
        severity_styles = {
            0: PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),  #Not classified
            1: PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  #Information
            2: PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  #Warning
            3: PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),  #Average
            4: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  #High
            5: PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")   #Disaster
        }
        
        #Цвета для статуса
        status_styles = {
            'PROBLEM': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            'RESOLVED': PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        }
        
        #Заголовки 
        headers = [
            "Time", "Severity", "Recovery Time", "Status", "HostGroup", 
            "Host", "Problem", "Ack", "Duration", "Tags DP", "Tags 2"
        ]
        
        # Ширина столбцов
        column_widths = [16, 15, 16, 13, 44, 37, 62, 5, 10, 11, 70]
        
        # Записываем заголовки
        for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border
            ws.column_dimensions[get_column_letter(col_num)].width = width
